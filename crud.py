from flask import Flask, render_template, request, redirect, url_for, send_file
import mysql.connector
from contextlib import closing
import io

# Reportes PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Reportes Excel
import openpyxl



# =====================
# inatalar la siguiente libreria = pip install reportlab openpyxl
# =====================

app = Flask(__name__)

def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="parrilla51"
    )

def obtener_categorias():
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("""
            SELECT * FROM categorias 
            WHERE nombre_categoria IN ('Bebidas', 'Platillos', 'Adicionales', 'Entradas', 'Acompañamientos')
        """)
        return cursor.fetchall()

# =====================
# RUTA PRINCIPAL → INICIO ADMIN
# =====================
@app.route("/")
def index():
    return redirect(url_for("inicioadmin"))

# =====================
# INICIO ADMIN PRINCIPAL
# =====================
@app.route("/inicioadmin")
def inicioadmin():
    return render_template("inicioadmin.html")

# =====================
# INVENTARIO: LISTAR PRODUCTOS, INSUMOS Y MESAS
# =====================
@app.route("/inventario")
def inventario():
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        # Productos
        cursor.execute("""
            SELECT p.id_producto, p.nombre, p.cantidad, p.descripcion, p.precio, 
                   p.imagen, c.nombre_categoria, p.cod_categoria
            FROM productos p
            JOIN categorias c ON p.cod_categoria = c.id_categoria
        """)
        productos = cursor.fetchall()

        # Insumos
        cursor.execute("""
            SELECT i.id_insumo, i.nombre, i.cantidad, i.precio, i.fecha_vencimiento, i.lote,
                   s.nombre_subcategoria, i.subcategoria_id
            FROM insumos i
            LEFT JOIN subcategorias_insumos s ON i.subcategoria_id = s.id_subcategoria
        """)
        insumos = cursor.fetchall()

        # Mesas
        cursor.execute("SELECT * FROM mesas")
        mesas = cursor.fetchall()

        # Categorías y subcategorías
        categorias = obtener_categorias()
        cursor.execute("SELECT * FROM subcategorias_insumos")
        subcategorias = cursor.fetchall()

    return render_template("inventario.html", 
                           productos=productos, 
                           insumos=insumos, 
                           mesas=mesas,
                           categorias=categorias, 
                           subcategorias=subcategorias)

# =====================
# CRUD PRODUCTOS
# =====================
@app.route("/producto/agregar", methods=["GET", "POST"])
def agregar_producto():
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            descripcion = request.form["descripcion"].strip()
            precio = float(str(request.form["precio"]).replace(",", "."))
            cod_categoria = int(request.form["cod_categoria"])
            imagen = request.form["imagen"].strip()
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
            cursor.execute("""
                INSERT INTO productos (nombre, cantidad, descripcion, precio, cod_categoria, imagen)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (nombre, cantidad, descripcion, precio, cod_categoria, imagen))
            conn.commit()
        return redirect(url_for("inventario"))

    categorias = obtener_categorias()
    return render_template("editar_producto.html", producto=None, categorias=categorias)


@app.route("/producto/editar/<int:id_producto>", methods=["GET", "POST"])
def editar_producto(id_producto):
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        if request.method == "POST":
            try:
                nombre = request.form["nombre"].strip()
                cantidad = int(request.form["cantidad"])
                descripcion = request.form["descripcion"].strip()
                precio = float(str(request.form["precio"]).replace(",", "."))
                cod_categoria = int(request.form["cod_categoria"])
                imagen = request.form["imagen"].strip()
            except (ValueError, KeyError):
                return "Datos inválidos", 400

            cursor.execute("""
                UPDATE productos 
                SET nombre=%s, cantidad=%s, descripcion=%s, precio=%s, cod_categoria=%s, imagen=%s
                WHERE id_producto=%s
            """, (nombre, cantidad, descripcion, precio, cod_categoria, imagen, id_producto))
            conn.commit()
            return redirect(url_for("inventario"))

        cursor.execute("SELECT * FROM productos WHERE id_producto=%s", (id_producto,))
        producto = cursor.fetchone()
        categorias = obtener_categorias()

    if not producto:
        return "Producto no encontrado", 404

    return render_template("editar_producto.html", producto=producto, categorias=categorias)

@app.route("/producto/eliminar/<int:id_producto>", methods=["POST"])
def eliminar_producto(id_producto):
    with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
        cursor.execute("DELETE FROM productos WHERE id_producto=%s", (id_producto,))
        conn.commit()
    return redirect(url_for("inventario"))

# =====================
# CRUD INSUMOS
# =====================
@app.route("/insumo/agregar", methods=["GET", "POST"])
def agregar_insumo():
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            precio = float(request.form["precio"])
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None
            lote = request.form.get("lote") or None
            subcategoria_id = int(request.form["subcategoria_id"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
            cursor.execute("""
                INSERT INTO insumos (nombre, cantidad, precio, fecha_vencimiento, lote, subcategoria_id)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (nombre, cantidad, precio, fecha_vencimiento, lote, subcategoria_id))
            conn.commit()
        return redirect(url_for("inventario"))

    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("SELECT * FROM subcategorias_insumos")
        subcategorias = cursor.fetchall()

    return render_template("editar_insumo.html", insumo=None, subcategorias=subcategorias)

@app.route("/insumo/editar/<int:id>", methods=["GET", "POST"])
def editar_insumo(id):
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        if request.method == "POST":
            try:
                nombre = request.form["nombre"].strip()
                cantidad = int(request.form["cantidad"])
                precio = float(request.form["precio"])
                fecha_vencimiento = request.form.get("fecha_vencimiento") or None
                lote = request.form.get("lote") or None
                subcategoria_id = int(request.form["subcategoria_id"])
            except (ValueError, KeyError):
                return "Datos inválidos", 400

            cursor.execute("""
                UPDATE insumos 
                SET nombre=%s, cantidad=%s, precio=%s, fecha_vencimiento=%s, lote=%s, subcategoria_id=%s
                WHERE id_insumo=%s
            """, (nombre, cantidad, precio, fecha_vencimiento, lote, subcategoria_id, id))
            conn.commit()
            return redirect(url_for("inventario"))

        cursor.execute("SELECT * FROM insumos WHERE id_insumo=%s", (id,))
        insumo = cursor.fetchone()
        cursor.execute("SELECT * FROM subcategorias_insumos")
        subcategorias = cursor.fetchall()

    if not insumo:
        return "Insumo no encontrado", 404

    return render_template("editar_insumo.html", insumo=insumo, subcategorias=subcategorias)

@app.route("/insumo/eliminar/<int:id>", methods=["POST"])
def eliminar_insumo(id):
    with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
        cursor.execute("DELETE FROM insumos WHERE id_insumo=%s", (id,))
        conn.commit()
    return redirect(url_for("inventario"))

# =====================
# CRUD MESAS
# =====================
@app.route("/mesa/agregar", methods=["GET", "POST"])
def agregar_mesa():
    if request.method == "POST":
        try:
            numero = int(request.form["numero"])
            capacidad = int(request.form["capacidad"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
            cursor.execute("""
                INSERT INTO mesas (numero, capacidad, estado)
                VALUES (%s, %s, 'Disponible')
            """, (numero, capacidad))
            conn.commit()
        return redirect(url_for("inventario"))

    return render_template("editar_mesa.html", mesa=None)


@app.route("/mesa/cambiar_estado/<int:id_mesa>")
def cambiar_estado(id_mesa):
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("SELECT estado FROM mesas WHERE id_mesa=%s", (id_mesa,))
        mesa = cursor.fetchone()
        if not mesa:
            return "Mesa no encontrada", 404

        nuevo_estado = "Disponible" if mesa["estado"] == "No disponible" else "No disponible"
        cursor.execute("UPDATE mesas SET estado=%s WHERE id_mesa=%s", (nuevo_estado, id_mesa))
        conn.commit()
    return redirect(url_for("inventario"))


@app.route("/mesa/eliminar/<int:id_mesa>", methods=["POST"])
def eliminar_mesa(id_mesa):
    with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
        cursor.execute("DELETE FROM mesas WHERE id_mesa=%s", (id_mesa,))
        conn.commit()
    return redirect(url_for("inventario"))

# =====================
# CONFIRMACIÓN DE ELIMINACIÓN
# =====================
@app.route("/confirmar_eliminacion/<string:tipo>/<int:item_id>")
def confirmar_eliminacion(tipo, item_id):
    if tipo == "producto":
        url = url_for("eliminar_producto", id_producto=item_id)
    elif tipo == "insumo":
        url = url_for("eliminar_insumo", id=item_id)
    elif tipo == "mesa":
        url = url_for("eliminar_mesa", id_mesa=item_id)
    else:
        return "Tipo no válido", 400

    return render_template("eliminar.html", tipo=tipo, item_id=item_id, url=url)


# =====================
# ROLES DE USUARIOS Y ESTADO
# =====================

@app.route("/asignarol", methods=["GET", "POST"])
def asignarol():
    if request.method == "POST":
        user_id = request.form["id_usuario"]
        nuevo_rol = request.form["rol"]
        nuevo_estado = request.form["estado"]

        with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
            cursor.execute("""
                UPDATE usuarios 
                SET rol = %s, estado = %s 
                WHERE id_usuario = %s
            """, (nuevo_rol, nuevo_estado, user_id))
            conn.commit()
        return redirect(url_for("asignarol"))

    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("SELECT * FROM usuarios")
        usuarios = cursor.fetchall()

    return render_template("asignarol.html", usuarios=usuarios)


# 🔹 Cambié el nombre de la función para evitar conflicto con cambiar_estado de mesas
@app.route("/usuario/cambiar_estado/<int:id_usuario>/<string:nuevo_estado>")
def cambiar_estado_usuario(id_usuario, nuevo_estado):
    with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
        cursor.execute("UPDATE usuarios SET estado = %s WHERE id_usuario = %s", (nuevo_estado, id_usuario))
        conn.commit()
    return redirect(url_for("asignarol"))


# 🔹 También renombré esta por claridad
@app.route("/usuario/cambiar_rol/<int:id_usuario>/<string:nuevo_rol>")
def cambiar_rol_usuario(id_usuario, nuevo_rol):
    with closing(get_db_connection()) as conn, closing(conn.cursor()) as cursor:
        cursor.execute("UPDATE usuarios SET rol = %s WHERE id_usuario = %s", (nuevo_rol, id_usuario))
        conn.commit()
    return redirect(url_for("asignarol"))


# CONSULTAR RESERVAS
@app.route("/consultar_reservas")
def consultar_reservas():
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("""
            SELECT 
                id_reserva,
                fecha,
                hora,
                cant_personas,
                estado,
                cod_mesa,
                telefono,
                id_usuario,
                nombre
            FROM reservas
            WHERE estado IN ('aceptada', 'cancelada')
            ORDER BY fecha ASC, hora ASC
        """)
        reservas = cursor.fetchall()

    return render_template("consultar_reservas.html", reservas=reservas)


@app.route("/consultaVentas")
def consultaVentas():
    pedidos_restaurante = []
    pedidos_domicilio = []
    try:
        with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
            # Pedidos en restaurante
            cursor.execute("""
                SELECT 
                    p.id_pedido,
                    u.nombre AS cliente,
                    p.fecha,
                    p.hora,
                    p.metodo_pago,
                    p.telefono,
                    p.total,
                    p.estado
                FROM pedidos p
                LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
                WHERE p.tipo_entrega = 'restaurante'
                  AND p.estado IN ('entregado','cancelado')
                ORDER BY p.fecha DESC, p.hora DESC
            """)
            pedidos_restaurante = cursor.fetchall()

            # Pedidos a domicilio
            cursor.execute("""
                SELECT 
                    p.id_pedido,
                    u.nombre AS cliente,
                    p.fecha,
                    p.hora,
                    p.metodo_pago,
                    p.telefono,
                    p.total,
                    p.estado
                FROM pedidos p
                LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
                WHERE p.tipo_entrega = 'domicilio'
                  AND p.estado IN ('entregado','cancelado')
                ORDER BY p.fecha DESC, p.hora DESC
            """)
            pedidos_domicilio = cursor.fetchall()

    except Exception as e:
        print("Error consultando la BD:", e)

    return render_template("consultaVentas.html",
                           pedidos_restaurante=pedidos_restaurante,
                           pedidos_domicilio=pedidos_domicilio)



# =====================
# CONSULTAR PRODUCTOS
# =====================
@app.route("/consulta_P")
def consulta_P():
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("""
            SELECT p.id_producto, p.nombre, p.cantidad, p.descripcion, p.precio, 
                   p.imagen, c.nombre_categoria
            FROM productos p
            JOIN categorias c ON p.cod_categoria = c.id_categoria
            ORDER BY p.nombre ASC
        """)
        productos = cursor.fetchall()

    return render_template("consulta_P.html", productos=productos)


# =====================
# CONSULTAR INSUMOS
# =====================
@app.route("/consulta_Y")
def consulta_Y():
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        cursor.execute("""
            SELECT i.id_insumo, i.nombre, i.cantidad, i.precio, i.fecha_vencimiento, i.lote,
                   s.nombre_subcategoria
            FROM insumos i
            LEFT JOIN subcategorias_insumos s ON i.subcategoria_id = s.id_subcategoria
            ORDER BY i.nombre ASC
        """)
        insumos = cursor.fetchall()

    return render_template("consulta_Y.html", insumos=insumos)



# ==================== REPORTES ====================

@app.route("/reportes")
def reportes():
    with closing(get_db_connection()) as conn, closing(conn.cursor(dictionary=True)) as cursor:
        # Pedidos restaurante
        cursor.execute("""
            SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, p.estado
            FROM pedidos p
            JOIN usuarios u ON p.cod_usuario = u.id_usuario
            WHERE p.tipo_entrega = 'restaurante'
        """)
        pedidos_restaurante = cursor.fetchall()

        # Pedidos domicilio
        cursor.execute("""
            SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, p.estado
            FROM pedidos p
            JOIN usuarios u ON p.cod_usuario = u.id_usuario
            WHERE p.tipo_entrega = 'domicilio'
        """)
        pedidos_domicilio = cursor.fetchall()

        # Reservas (usar la vista)
        cursor.execute("SELECT * FROM vista_reservas_mesas")
        reservas = cursor.fetchall()

        # Inventario bajo (solo insumos, no productos)
        cursor.execute("SELECT * FROM vista_insumos_stock_bajo")
        stock_bajo = cursor.fetchall()

    return render_template(
        "reportes.html",
        pedidos_restaurante=pedidos_restaurante or [],
        pedidos_domicilio=pedidos_domicilio or [],
        reservas=reservas or [],
        stock_bajo=stock_bajo or []
    )


# ==================== REPORTES PDF ====================

@app.route("/reportes/pdf")
def reportes_pdf():
    with closing(get_db_connection()) as conn:
        with closing(conn.cursor(dictionary=True)) as cursor:
            # Pedidos en Restaurante
            cursor.execute("""
                SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, p.estado
                FROM pedidos p
                INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
                WHERE p.tipo_entrega = 'restaurante'
            """)
            pedidos_restaurante = cursor.fetchall()

            # Pedidos a Domicilio
            cursor.execute("""
                SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, p.estado
                FROM pedidos p
                INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
                WHERE p.tipo_entrega = 'domicilio'
            """)
            pedidos_domicilio = cursor.fetchall()

            # Reservas (desde vista con mesa y capacidad)
            cursor.execute("SELECT * FROM vista_reservas_mesas")
            reservas = cursor.fetchall()

            # Inventario bajo
            cursor.execute("SELECT * FROM insumos WHERE cantidad < 5")
            inventario = cursor.fetchall()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # ---- Pedidos en Restaurante ----
    elements.append(Paragraph("Pedidos en Restaurante", styles['Heading2']))
    data = [["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"]]
    if pedidos_restaurante:
        for p in pedidos_restaurante:
            data.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}", str(p['fecha']),
                         str(p['hora']), f"${p['total']:.2f}", p['estado']])
    else:
        data.append(["-", "No hay pedidos en restaurante", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # ---- Pedidos a Domicilio ----
    elements.append(Paragraph("Pedidos a Domicilio", styles['Heading2']))
    data = [["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"]]
    if pedidos_domicilio:
        for p in pedidos_domicilio:
            data.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}", str(p['fecha']),
                         str(p['hora']), f"${p['total']:.2f}", p['estado']])
    else:
        data.append(["-", "No hay pedidos a domicilio", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # ---- Reservas ----
    elements.append(Paragraph("Reservas", styles['Heading2']))
    data = [["ID", "Fecha", "Hora", "Personas", "Estado", "Mesa", "Capacidad"]]
    if reservas:
        for r in reservas:
            data.append([r['id_reserva'], str(r['fecha']), str(r['hora']), r['cant_personas'],
                         r['estado'], r['mesa'], r['capacidad']])
    else:
        data.append(["-", "No hay reservas registradas", "-", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # ---- Inventario ----
    elements.append(Paragraph("Inventario Bajo", styles['Heading2']))
    data = [["ID Insumo", "Nombre", "Cantidad", "Precio"]]
    if inventario:
        for i in inventario:
            data.append([i['id_insumo'], i['nombre'], i['cantidad'], f"${i['precio']:.2f}"])
    else:
        data.append(["-", "No hay insumos con stock bajo", "-", "-"])
    elements.append(Table(data))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="reporte_general.pdf", mimetype="application/pdf")


# ==================== REPORTES EXCEL ====================

@app.route("/reportes/excel")
def reportes_excel():
    with closing(get_db_connection()) as conn:
        with closing(conn.cursor(dictionary=True)) as cursor:
            cursor.execute("""
                SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, p.estado
                FROM pedidos p
                INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
                WHERE p.tipo_entrega = 'restaurante'
            """)
            pedidos_restaurante = cursor.fetchall()

            cursor.execute("""
                SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, p.estado
                FROM pedidos p
                INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
                WHERE p.tipo_entrega = 'domicilio'
            """)
            pedidos_domicilio = cursor.fetchall()

            cursor.execute("SELECT * FROM vista_reservas_mesas")
            reservas = cursor.fetchall()

            cursor.execute("SELECT * FROM insumos WHERE cantidad < 5")
            inventario = cursor.fetchall()

    wb = openpyxl.Workbook()

    # ---- Pedidos Restaurante ----
    ws = wb.active
    ws.title = "Pedidos Restaurante"
    ws.append(["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"])
    if pedidos_restaurante:
        for p in pedidos_restaurante:
            ws.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}", str(p['fecha']),
                       str(p['hora']), p['total'], p['estado']])
    else:
        ws.append(["-", "No hay pedidos en restaurante", "-", "-", "-", "-"])

    # ---- Pedidos Domicilio ----
    ws = wb.create_sheet("Pedidos Domicilio")
    ws.append(["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"])
    if pedidos_domicilio:
        for p in pedidos_domicilio:
            ws.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}", str(p['fecha']),
                       str(p['hora']), p['total'], p['estado']])
    else:
        ws.append(["-", "No hay pedidos a domicilio", "-", "-", "-", "-"])

    # ---- Reservas ----
    ws = wb.create_sheet("Reservas")
    ws.append(["ID Reserva", "Fecha", "Hora", "Personas", "Estado", "Mesa", "Capacidad"])
    if reservas:
        for r in reservas:
            ws.append([r['id_reserva'], str(r['fecha']), str(r['hora']), r['cant_personas'],
                       r['estado'], r['mesa'], r['capacidad']])
    else:
        ws.append(["-", "No hay reservas registradas", "-", "-", "-", "-", "-"])

    # ---- Inventario Bajo ----
    ws = wb.create_sheet("Inventario Bajo")
    ws.append(["ID Insumo", "Nombre", "Cantidad", "Precio"])
    if inventario:
        for i in inventario:
            ws.append([i['id_insumo'], i['nombre'], i['cantidad'], i['precio']])
    else:
        ws.append(["-", "No hay insumos con stock bajo", "-", "-"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="reporte_general.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =====================
# MAIN
# =====================
if __name__ == "__main__":
    app.run(debug=True)
    